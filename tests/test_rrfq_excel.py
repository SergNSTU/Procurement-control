from __future__ import annotations

import unittest
from decimal import Decimal

import openpyxl

from rrfq_app.lcsc import LcscClient
from rrfq_app.models import SupplierQuote
from rrfq_app.resolver import resolve_items
from rrfq_app.rrfq_excel import fill_rrfq, load_visible_items


class RrfqExcelTest(unittest.TestCase):
    def test_load_visible_items_skips_hidden_rows(self):
        with _tmpdir() as tmp_path:
            workbook_path = tmp_path / "rfq.xlsx"
            _make_workbook(workbook_path)

            items = load_visible_items(workbook_path)

        self.assertEqual([item.row for item in items], [49])
        self.assertEqual(items[0].value, "ABC123")
        self.assertEqual(items[0].requested_mfg, "OriginalMfg")

    def test_fill_rrfq_writes_best_price_and_preserves_lead_time(self):
        with _tmpdir() as tmp_path:
            workbook_path = tmp_path / "rfq.xlsx"
            output_path = tmp_path / "out.xlsx"
            _make_workbook(workbook_path)
            quotes = [
                SupplierQuote(
                    row=49,
                    value="ABC123",
                    offer_pn="ABC123-G",
                    offer_mfg="OfferMfg",
                    supplier="LCSC",
                    currency="USD",
                    unit_price=Decimal("2.50"),
                    pack_qty=50,
                ),
                SupplierQuote(
                    row=49,
                    value="ABC123",
                    offer_pn="ABC123",
                    offer_mfg="RusOffer",
                    supplier="RF",
                    currency="RUB",
                    unit_price=Decimal("244"),
                ),
            ]

            applied = fill_rrfq(workbook_path, output_path, quotes, usd_rub=Decimal("100"))

            result = openpyxl.load_workbook(output_path, data_only=False)
            sheet = result["BOM"]

        self.assertEqual(len(applied), 1)
        self.assertEqual(applied[0].quote.supplier, "RF")
        self.assertEqual(sheet["I49"].value, "ABC123")
        self.assertEqual(sheet["O49"].value, "RusOffer")
        self.assertEqual(sheet["R49"].value, 2.0)
        self.assertEqual(sheet["S49"].value, "=R49*Q49")
        self.assertEqual(sheet["T49"].value, "do not touch")
        self.assertIs(sheet.row_dimensions[50].hidden, True)

    def test_estimated_price_gets_red_fill_and_comment(self):
        with _tmpdir() as tmp_path:
            workbook_path = tmp_path / "rfq.xlsx"
            output_path = tmp_path / "out.xlsx"
            _make_workbook(workbook_path)
            quotes = [
                SupplierQuote(
                    row=49,
                    value="ABC123",
                    offer_pn="ABC123",
                    offer_mfg="OriginalMfg",
                    supplier="Estimate",
                    currency="USD",
                    unit_price=Decimal("1.23"),
                    is_estimated=True,
                    reason="No direct price.",
                )
            ]

            fill_rrfq(workbook_path, output_path, quotes)
            result = openpyxl.load_workbook(output_path, data_only=False)
            sheet = result["BOM"]

        self.assertEqual(sheet["R49"].value, 1.23)
        self.assertEqual(sheet["R49"].fill.fgColor.rgb, "FFFFC7CE")
        self.assertIn("No direct price", sheet["R49"].comment.text)
        self.assertEqual(sheet["T49"].value, "do not touch")

    def test_direct_price_clears_previous_estimate_marker(self):
        with _tmpdir() as tmp_path:
            workbook_path = tmp_path / "rfq.xlsx"
            estimated_path = tmp_path / "estimated.xlsx"
            direct_path = tmp_path / "direct.xlsx"
            _make_workbook(workbook_path)
            fill_rrfq(
                workbook_path,
                estimated_path,
                [
                    SupplierQuote(
                        row=49,
                        value="ABC123",
                        offer_pn="ABC123",
                        offer_mfg="OriginalMfg",
                        supplier="Estimate",
                        currency="USD",
                        unit_price=Decimal("1.23"),
                        is_estimated=True,
                        reason="No direct price.",
                    )
                ],
            )
            fill_rrfq(
                estimated_path,
                direct_path,
                [
                    SupplierQuote(
                        row=49,
                        value="ABC123",
                        offer_pn="ABC123",
                        offer_mfg="OriginalMfg",
                        supplier="Direct",
                        currency="USD",
                        unit_price=Decimal("1.11"),
                        confidence=Decimal("0.95"),
                    )
                ],
            )
            result = openpyxl.load_workbook(direct_path, data_only=False)
            sheet = result["BOM"]

        self.assertEqual(sheet["R49"].value, 1.11)
        self.assertNotEqual(str(sheet["R49"].fill.fgColor.rgb), "FFFFC7CE")
        self.assertIsNone(sheet["R49"].comment)

    def test_resolver_without_api_credentials_estimates_visible_items(self):
        with _tmpdir() as tmp_path:
            workbook_path = tmp_path / "rfq.xlsx"
            _make_workbook(workbook_path)
            items = load_visible_items(workbook_path)

        decisions = resolve_items(items, LcscClient(None))

        self.assertEqual(len(decisions), 1)
        self.assertEqual(decisions[0].status, "estimated")
        self.assertTrue(decisions[0].quote.quote.is_estimated)


def _make_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "BOM"

    headers = {
        "A48": "#",
        "B48": "Process",
        "C48": "Eng. name",
        "F48": "Обозначение/величина ",
        "I48": "Каталожный номер",
        "K48": "QTY in order",
        "M48": "Body",
        "N48": "Производитель из РФ",
        "O48": "Производитель из КНР",
        "P48": "Кол-во в упаковке ",
        "Q48": "Кол-во закупки",
        "R48": "Цена за единицу",
        "S48": "Итоговая цена",
        "T48": "Срок доставки (день)",
    }
    for cell, value in headers.items():
        sheet[cell] = value

    sheet["A49"] = 1
    sheet["B49"] = "SMT"
    sheet["C49"] = "IC"
    sheet["F49"] = "ABC123"
    sheet["K49"] = 100
    sheet["M49"] = "SOIC-8"
    sheet["N49"] = "OriginalMfg"
    sheet["Q49"] = 150
    sheet["S49"] = "=R49*Q49"
    sheet["T49"] = "do not touch"

    sheet["A50"] = 2
    sheet["F50"] = "HIDDEN"
    sheet["Q50"] = 100
    sheet["S50"] = "=R50*Q50"
    sheet.row_dimensions[50].hidden = True

    workbook.save(path)


class _tmpdir:
    def __enter__(self):
        import shutil
        import uuid
        from pathlib import Path

        self.path = Path.cwd() / ".test_tmp" / uuid.uuid4().hex
        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir(parents=True)
        return self.path

    def __exit__(self, exc_type, exc, tb):
        import shutil

        shutil.rmtree(self.path, ignore_errors=True)
        return False


if __name__ == "__main__":
    unittest.main()
