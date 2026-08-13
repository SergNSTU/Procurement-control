from __future__ import annotations

from copy import copy
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import NormalizedQuote, RrfqColumns, RrfqItem, SupplierQuote
from .pricing import choose_best_quote, decimal_from_cell


HEADER_ALIASES = {
    "process": {"process", "тех. операция"},
    "category": {"eng. name", "наименование англ."},
    "value": {"value", "обозначение/величина"},
    "pn": {"pn", "каталожный номер"},
    "mfg_russia": {"mfg from russia", "производитель из рф"},
    "mfg_china": {"mfg from china", "производитель из кнр"},
    "qty_in_order": {"qty in order", "кол-во всего"},
    "body": {"body", "корпус"},
    "pack_qty": {"qty in packing", "кол-во в упаковке"},
    "qty_to_buy": {"qty to buy/pcs", "кол-во закупки"},
    "unit_price": {"unit price", "цена за единицу"},
    "total_price": {"total price", "итоговая цена"},
    "lead_time": {"lead time (days)", "срок доставки (день)"},
}

ESTIMATED_PRICE_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")


class RrfqTemplateError(ValueError):
    pass


def load_visible_items(path: str | Path) -> list[RrfqItem]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    values_workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet, header_row, columns = find_rrfq_table(workbook)
    values_worksheet = values_workbook[worksheet.title]

    items: list[RrfqItem] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        if worksheet.row_dimensions[row].hidden:
            continue

        item_no = worksheet.cell(row, columns.item_no).value
        value = worksheet.cell(row, columns.value).value
        if not isinstance(item_no, (int, float)) or not value:
            continue

        items.append(
            RrfqItem(
                sheet_name=worksheet.title,
                row=row,
                item_no=int(item_no),
                value=str(value).strip(),
                requested_mfg=_cell_text(worksheet.cell(row, columns.mfg_russia)),
                current_pn=_cell_text(worksheet.cell(row, columns.pn)),
                current_offer_mfg=_cell_text(worksheet.cell(row, columns.mfg_china)),
                qty_to_buy=decimal_from_cell(values_worksheet.cell(row, columns.qty_to_buy).value),
                process=_optional_cell_text(worksheet, row, columns.process),
                category=_optional_cell_text(worksheet, row, columns.category),
                body=_optional_cell_text(worksheet, row, columns.body),
                qty_in_order=_optional_cell_decimal(values_worksheet, row, columns.qty_in_order),
            )
        )

    return items


def fill_rrfq(
    input_path: str | Path,
    output_path: str | Path,
    quotes: list[SupplierQuote],
    *,
    usd_rub: Decimal | None = None,
    usd_cny: Decimal | None = None,
) -> list[NormalizedQuote]:
    workbook = openpyxl.load_workbook(input_path, data_only=False)
    values_workbook = openpyxl.load_workbook(input_path, data_only=True)
    worksheet, header_row, columns = find_rrfq_table(workbook)
    values_worksheet = values_workbook[worksheet.title]

    applied: list[NormalizedQuote] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        if worksheet.row_dimensions[row].hidden:
            continue

        item_no = worksheet.cell(row, columns.item_no).value
        value = worksheet.cell(row, columns.value).value
        if not isinstance(item_no, (int, float)) or not value:
            continue

        item = RrfqItem(
            sheet_name=worksheet.title,
            row=row,
            item_no=int(item_no),
            value=str(value).strip(),
            requested_mfg=_cell_text(worksheet.cell(row, columns.mfg_russia)),
            current_pn=_cell_text(worksheet.cell(row, columns.pn)),
            current_offer_mfg=_cell_text(worksheet.cell(row, columns.mfg_china)),
            qty_to_buy=decimal_from_cell(values_worksheet.cell(row, columns.qty_to_buy).value),
            process=_optional_cell_text(worksheet, row, columns.process),
            category=_optional_cell_text(worksheet, row, columns.category),
            body=_optional_cell_text(worksheet, row, columns.body),
            qty_in_order=_optional_cell_decimal(values_worksheet, row, columns.qty_in_order),
        )
        best = choose_best_quote(item, quotes, usd_rub=usd_rub, usd_cny=usd_cny)
        if not best:
            continue

        _write_offer(worksheet, row, columns, best)
        applied.append(best)

    _force_recalculation(workbook)
    workbook.save(output_path)
    return applied


def find_rrfq_table(workbook: openpyxl.Workbook) -> tuple[Worksheet, int, RrfqColumns]:
    for worksheet in workbook.worksheets:
        for row in range(1, worksheet.max_row + 1):
            mapping = _header_mapping(worksheet, row)
            required = {
                "value",
                "pn",
                "mfg_russia",
                "mfg_china",
                "pack_qty",
                "qty_to_buy",
                "unit_price",
                "total_price",
                "lead_time",
            }
            if required.issubset(mapping.keys()):
                return (
                    worksheet,
                    row,
                    RrfqColumns(
                        item_no=1,
                        value=mapping["value"],
                        pn=mapping["pn"],
                        mfg_russia=mapping["mfg_russia"],
                        mfg_china=mapping["mfg_china"],
                        pack_qty=mapping["pack_qty"],
                        qty_to_buy=mapping["qty_to_buy"],
                        unit_price=mapping["unit_price"],
                        total_price=mapping["total_price"],
                        lead_time=mapping["lead_time"],
                        process=mapping.get("process"),
                        category=mapping.get("category"),
                        body=mapping.get("body"),
                        qty_in_order=mapping.get("qty_in_order"),
                    ),
                )

    raise RrfqTemplateError("Could not find an RRFQ table header.")


def _header_mapping(worksheet: Worksheet, row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in worksheet[row]:
        normalized = _normalize_header(cell.value)
        if not normalized:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapping[field] = cell.column
    return mapping


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _cell_text(cell: Cell) -> str | None:
    value = cell.value
    if value in (None, ""):
        return None
    return str(value).strip()


def _optional_cell_text(worksheet: Worksheet, row: int, column: int | None) -> str | None:
    if column is None:
        return None
    return _cell_text(worksheet.cell(row, column))


def _optional_cell_decimal(worksheet: Worksheet, row: int, column: int | None) -> Decimal | None:
    if column is None:
        return None
    return decimal_from_cell(worksheet.cell(row, column).value)


def _write_offer(
    worksheet: Worksheet,
    row: int,
    columns: RrfqColumns,
    best: NormalizedQuote,
) -> None:
    quote = best.quote

    if quote.offer_pn:
        _copy_style_and_write(worksheet.cell(row, columns.pn), quote.offer_pn)
    if quote.offer_mfg:
        _copy_style_and_write(worksheet.cell(row, columns.mfg_china), quote.offer_mfg)
    if quote.pack_qty:
        _copy_style_and_write(worksheet.cell(row, columns.pack_qty), quote.pack_qty)

    price_cell = worksheet.cell(row, columns.unit_price)
    _copy_style_and_write(price_cell, float(best.unit_price_usd))
    if quote.is_estimated:
        price_cell.fill = copy(ESTIMATED_PRICE_FILL)
        price_cell.comment = Comment(
            quote.reason or "Estimated price; no direct supplier quote found.",
            "RRFQ Pricer",
        )
    else:
        if _is_estimated_price_fill(price_cell):
            price_cell.fill = PatternFill(fill_type=None)
        if price_cell.comment and price_cell.comment.author == "RRFQ Pricer":
            price_cell.comment = None

    # Lead time is intentionally left untouched for pre-calculation.
    total_cell = worksheet.cell(row, columns.total_price)
    if not isinstance(total_cell.value, str) or not total_cell.value.startswith("="):
        total_cell.value = f"={get_column_letter(columns.unit_price)}{row}*{get_column_letter(columns.qty_to_buy)}{row}"


def _copy_style_and_write(cell: Cell, value: object) -> None:
    style = copy(cell._style)
    number_format = cell.number_format
    alignment = copy(cell.alignment)
    cell.value = value
    cell._style = style
    cell.number_format = number_format
    cell.alignment = alignment


def _force_recalculation(workbook: openpyxl.Workbook) -> None:
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def _is_estimated_price_fill(cell: Cell) -> bool:
    return str(cell.fill.fgColor.rgb) == "FFFFC7CE"
