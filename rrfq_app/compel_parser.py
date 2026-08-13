from __future__ import annotations

import html
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class CompelRow:
    number: int | None
    source_name: str
    matched_part: str
    manufacturer: str
    package: str
    stock_count: int | None
    lead_time: str
    quantity: int | None
    unit_price: float | None
    currency: str
    total: float | None
    total_currency: str


@dataclass(frozen=True)
class CompelSummary:
    total_rows: int
    priced_rows: int
    missing_rows: list[int]
    selected_total_usd: float
    longest_lead_days: int | None
    longest_lead_time: str
    cheap_total_usd: float | None
    cheap_lead_time: str
    optimal_total_usd: float | None
    optimal_lead_time: str


HEADERS = [
    "N",
    "Исходное наименование",
    "Подобранный товар",
    "Производитель",
    "Срок поставки",
    "Количество",
    "Цена за шт",
    "Валюта",
    "Сумма",
    "Валюта суммы",
]


def parse_compel_html(path: str | Path) -> tuple[list[CompelRow], CompelSummary]:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    rows = _parse_rows(text)
    summary = _parse_summary(text, rows)
    return rows, summary


def write_compel_xlsx(path: str | Path, rows: list[CompelRow], summary: CompelSummary) -> None:
    workbook = openpyxl.Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Сводка"
    detail_ws = workbook.create_sheet("Позиции")

    _write_summary(summary_ws, summary)
    _write_details(detail_ws, rows)

    workbook.save(path)


def compel_row_json(row: CompelRow) -> dict:
    return {
        "number": row.number,
        "source_name": row.source_name,
        "matched_part": row.matched_part,
        "manufacturer": row.manufacturer,
        "package": row.package,
        "stock_count": row.stock_count,
        "lead_time": row.lead_time,
        "quantity": row.quantity,
        "unit_price": row.unit_price,
        "currency": row.currency,
        "total": row.total,
        "total_currency": row.total_currency,
    }


def compel_summary_json(summary: CompelSummary) -> dict:
    return {
        "total_rows": summary.total_rows,
        "priced_rows": summary.priced_rows,
        "missing_rows": summary.missing_rows,
        "selected_total_usd": round(summary.selected_total_usd, 6),
        "longest_lead_days": summary.longest_lead_days,
        "longest_lead_time": summary.longest_lead_time,
        "cheap_total_usd": summary.cheap_total_usd,
        "cheap_lead_time": summary.cheap_lead_time,
        "optimal_total_usd": summary.optimal_total_usd,
        "optimal_lead_time": summary.optimal_lead_time,
    }


def _parse_rows(text: str) -> list[CompelRow]:
    chunks = re.split(r'(?=<tr data-is="ac-row" class="ac-row)', text)
    rows: list[CompelRow] = []

    for chunk in chunks:
        if not chunk.startswith('<tr data-is="ac-row"'):
            continue

        number = _to_int(_clean(_first(r'<div class="ac-row-num">(.*?)</div>', chunk)))
        source_name = _clean(_first(r'<span[^>]*class="ac-row-name[^>]*>(.*?)</span>', chunk))
        matched_text = _clean(_first(r'<a[^>]*class="ac-item-pn-link"[^>]*>(.*?)</a>', chunk))
        matched_part, manufacturer = _split_matched_part(matched_text)
        package = _clean(_first(r'<span[^>]*class="ac-part-mpq"[^>]*>(.*?)</span>', chunk))
        stock_count = _to_int(_clean(_first(r'<span[^>]*class="ac-item-stocks-key"[^>]*>(.*?)</span>', chunk)))
        lead_time = _clean(_first(r'<span[^>]*class="ac-item-offer-dlv"[^>]*>(.*?)</span>', chunk))
        quantity = _to_int(re.sub(r"\s*x\s*$", "", _clean(_first(r'<span[^>]*class="ac-item-offer-qty"[^>]*>(.*?)</span>', chunk))))

        unit_price = None
        currency = ""
        unit_match = re.search(
            r'<span class="ac-item-offer-price">.*?<price2\s+val="([^"]+)"\s+cur="([^"]+)"',
            chunk,
            re.S,
        )
        if unit_match:
            unit_price = _to_float(unit_match.group(1))
            currency = unit_match.group(2)

        total = None
        total_currency = ""
        total_block = _first(r'<div class="ac-row-total-sum[^"]*"[^>]*>(.*?)</div>', chunk)
        if total_block:
            total_match = re.search(r'<price2\s+val="([^"]+)"\s+cur="([^"]+)"', total_block, re.S)
            if total_match:
                total = _to_float(total_match.group(1))
                total_currency = total_match.group(2)

        rows.append(
            CompelRow(
                number=number,
                source_name=source_name,
                matched_part=matched_part,
                manufacturer=manufacturer,
                package=package,
                stock_count=stock_count,
                lead_time=lead_time,
                quantity=quantity,
                unit_price=unit_price,
                currency=currency,
                total=total,
                total_currency=total_currency,
            )
        )

    return rows


def _parse_summary(text: str, rows: list[CompelRow]) -> CompelSummary:
    priced_rows = [row for row in rows if row.total is not None]
    missing_rows = [row.number for row in rows if row.total is None and row.number is not None]
    selected_total_usd = sum(row.total or 0 for row in rows)
    lead_days = [(days, row.lead_time) for row in rows if (days := _lead_days(row.lead_time)) is not None]
    longest_lead_days, longest_lead_time = max(lead_days, default=(None, ""), key=lambda item: item[0])

    cheap_total_usd = None
    cheap_lead_time = ""
    cheap_match = re.search(
        r'<div data-set="cheap"[^>]*>.*?<price2\s+val="([^"]+)"\s+cur="USD".*?'
        r'<span class="bom-settings-offer-type-dlv">([^<]+)</span>',
        text,
        re.S,
    )
    if cheap_match:
        cheap_total_usd = _to_float(cheap_match.group(1))
        cheap_lead_time = _clean(cheap_match.group(2))

    optimal_total_usd = None
    optimal_lead_time = ""
    optimal_match = re.search(
        r'<div data-set="optimal"[^>]*>.*?<price2\s+val="([^"]+)"\s+cur="USD".*?'
        r'<span class="bom-settings-offer-type-dlv">([^<]+)</span>',
        text,
        re.S,
    )
    if optimal_match:
        optimal_total_usd = _to_float(optimal_match.group(1))
        optimal_lead_time = _clean(optimal_match.group(2))

    return CompelSummary(
        total_rows=len(rows),
        priced_rows=len(priced_rows),
        missing_rows=missing_rows,
        selected_total_usd=selected_total_usd,
        longest_lead_days=longest_lead_days,
        longest_lead_time=longest_lead_time,
        cheap_total_usd=cheap_total_usd,
        cheap_lead_time=cheap_lead_time,
        optimal_total_usd=optimal_total_usd,
        optimal_lead_time=optimal_lead_time,
    )


def _write_summary(ws, summary: CompelSummary) -> None:
    ws.append(["Расчет SDS Compel"])
    ws.merge_cells("A1:D1")
    ws.append([])
    ws.append(["Показатель", "USD", "Комментарий", ""])
    ws.append(["Итого строк с подобранной ценой", summary.selected_total_usd, "Сумма по листу Позиции", ""])
    ws.append(["Итог сайта: оптимизировано по цене", summary.cheap_total_usd, summary.cheap_lead_time, ""])
    ws.append(["Итог сайта: оптимизировано по сроку", summary.optimal_total_usd, summary.optimal_lead_time, ""])
    ws.append(["Всего строк", summary.total_rows, "", ""])
    ws.append(["Строк с ценой", summary.priced_rows, "", ""])
    ws.append(["Строк без подобранного предложения", len(summary.missing_rows), ", ".join(map(str, summary.missing_rows)), ""])
    ws.append(["Самый долгий срок по позициям", summary.longest_lead_days, summary.longest_lead_time, ""])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    subheader_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    for cell in ws[3]:
        cell.fill = subheader_fill
        cell.font = Font(bold=True)
    for row in range(4, 11):
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).number_format = "#,##0.00"

    widths = {"A": 34, "B": 16, "C": 38, "D": 8}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_details(ws, rows: list[CompelRow]) -> None:
    ws.append(HEADERS)
    for row in rows:
        ws.append(
            [
                row.number,
                row.source_name,
                row.matched_part,
                row.manufacturer,
                row.lead_time,
                row.quantity,
                row.unit_price,
                row.currency,
                row.total,
                row.total_currency,
            ]
        )

    total_row = len(rows) + 2
    ws.cell(total_row, 2).value = "Итого"
    ws.cell(total_row, 5).value = "Самый долгий срок"
    ws.cell(total_row, 6).value = _max_lead_time(rows)
    ws.cell(total_row, 9).value = sum(row.total or 0 for row in rows)
    ws.cell(total_row, 10).value = "USD"
    for cell in ws[total_row]:
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    number_formats = {
        "A": "0",
        "F": "0",
        "G": "0.00000",
        "I": "#,##0.00",
    }
    for col, fmt in number_formats.items():
        for cell in ws[col][1:]:
            cell.number_format = fmt

    widths = {
        "A": 6,
        "B": 28,
        "C": 32,
        "D": 18,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 10,
        "I": 14,
        "J": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, re.S)
    return match.group(1) if match else ""


def _clean(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    value = html.unescape(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def _split_matched_part(value: str) -> tuple[str, str]:
    match = re.fullmatch(r"(.+?)\s*\(([^()]*)\)", value.strip())
    if not match:
        return value, ""
    return match.group(1).strip(), match.group(2).strip()


def _lead_days(value: str) -> int | None:
    match = re.search(r"\d+", value or "")
    return int(match.group(0)) if match else None


def _max_lead_time(rows: list[CompelRow]) -> str:
    values = [(days, row.lead_time) for row in rows if (days := _lead_days(row.lead_time)) is not None]
    return max(values, default=(0, ""), key=lambda item: item[0])[1]


def _to_int(value: str) -> int | None:
    if value in ("", None):
        return None
    value = str(value).replace(" ", "")
    if not re.fullmatch(r"-?\d+", value):
        return None
    return int(value)


def _to_float(value: str) -> float | None:
    if value in ("", None):
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None
